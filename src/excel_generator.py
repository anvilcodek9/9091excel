"""Excel file generator for Logen delivery format."""

import os
from datetime import datetime
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .exceptions import ExcelGenerationError

# 로젠양식: A~O 15열 (D,E / H,I / K 빈칸)
# A:수하인명 B:운송장번호 C:수하인주소1 D,E:빈칸 F:수하인핸드폰번호 G:택배수량
# H,I:빈칸 J:품목명 K:빈칸 L:배송메세지 M:보내는분 N:연락처 O:주소
LOGEN_HEADERS = [
    "수하인명",           # A
    "운송장번호(로젠택배)",  # B
    "수하인주소1",         # C
    "", "",                # D, E 빈칸
    "수하인핸드폰번호",     # F
    "택배수량",            # G
    "", "",                # H, I 빈칸
    "품목명",              # J
    "",                   # K 빈칸
    "배송메세지 (도착일)",  # L
    "보내는 분",           # M
    "연락처",              # N
    "주소",                # O
]
LOGEN_SHEET_NAME = "엑셀파일첫행-제목있음(주소1통합)"


def _adjust_column_widths(ws) -> None:
    """각 열의 최대 내용 길이에 맞춰 열 너비 조절."""
    for col_idx in range(1, len(LOGEN_HEADERS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is not None:
                s = str(val)
                # 한글 등은 2자로 처리 (대략적)
                try:
                    display_len = sum(2 if ord(c) > 127 else 1 for c in s)
                except (TypeError, ValueError):
                    display_len = len(s)
                max_len = max(max_len, display_len)
        if max_len > 0:
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


class LogenExcelGenerator:
    """Generator for Logen delivery Excel files."""
    
    @staticmethod
    def generate_excel(data: List[Dict], output_path: str) -> str:
        """
        Generate Logen shipping Excel file.
        
        Creates an Excel workbook in Logen format (로젠양식.xls) with 15 columns.
        
        Args:
            data: List of transformed order dictionaries with keys:
                - receiver_name: 수하인명
                - address1, address2 또는 full_address: 수하인주소1 한 칸에 통합 저장
                - receiver_tel: 수하인전화번호
                - product_name: 품목명 (옵션정보 우선)
                - delivery_memo: 배송메세지 (도착일)
                - sender_name: 보내는 분 (구매자명)
                - sender_tel: 보내는 분 연락처 (구매자 연락처)
            output_path: Path for output Excel file
            
        Returns:
            str: Path to generated file
            
        Raises:
            ExcelGenerationError: When file creation fails due to file system
                errors or other issues
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = LOGEN_SHEET_NAME
            
            # Create header row (로젠양식 A~O 15열)
            ws.append(LOGEN_HEADERS)
            
            # Create data rows (D,E / H,I / K 빈칸)
            for order in data:
                address1 = order.get("address1", "")
                address2 = order.get("address2", "")
                if not address1 and not address2:
                    full = order.get("full_address", "")
                    address1, address2 = (full, "") if full else ("", "")
                full_address = f"{address1} {address2}".strip() if (address1 or address2) else ""
                receiver_tel = order.get("receiver_tel", "")
                sender_name = order.get("sender_name") or order.get("buyer_name") or ""
                sender_tel = order.get("sender_tel") or order.get("buyer_tel") or ""
                sender_addr = order.get("sender_address", "")
                
                row = [
                    order.get("receiver_name", ""),      # A
                    "",                                  # B 운송장번호
                    full_address,                        # C 수하인주소1
                    "", "",                              # D, E 빈칸
                    receiver_tel,                         # F 수하인핸드폰번호
                    1,                                    # G 택배수량
                    "", "",                              # H, I 빈칸
                    order.get("product_name", ""),       # J 품목명 (옵션정보)
                    "",                                  # K 빈칸
                    order.get("delivery_memo", ""),      # L 배송메세지
                    sender_name,                         # M 보내는분 (구매자명)
                    sender_tel,                          # N 연락처 (구매자 연락처)
                    sender_addr,                         # O 주소 (보내는 분 주소, 선택)
                ]
                ws.append(row)
            
            # 열 너비 조절 (입력 글자 다 보이게)
            _adjust_column_widths(ws)

            # 파일명이 이미 존재하면 충돌을 피하기 위해 임시 값을 붙여서 저장
            final_path = output_path
            if os.path.exists(final_path):
                base, ext = os.path.splitext(final_path)
                # 시각 기반 임시값(HHMMSS) 우선 추가
                timestamp = datetime.now().strftime("%H%M%S")
                candidate = f"{base}_{timestamp}{ext}"
                counter = 1
                # 혹시 같은 시각에 여러 번 생성하는 경우를 대비해 카운터까지 추가
                while os.path.exists(candidate):
                    candidate = f"{base}_{timestamp}_{counter}{ext}"
                    counter += 1
                final_path = candidate

            wb.save(final_path)
            return final_path
            
        except Exception as e:
            raise ExcelGenerationError(
                message=f"Failed to generate Excel file",
                file_path=output_path,
                underlying_error=e
            )
