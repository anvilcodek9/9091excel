"""Utility functions for the Naver Smart Store - Logen integration."""

from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import load_workbook

# 로젠양식 헤더 (15열) - excel_generator와 동일
LOGEN_EXCEL_HEADERS = [
    "수하인명",
    "운송장번호(로젠택배)",
    "수하인주소1",
    "수하인주소2",
    "수하인전화번호",
    "수하인핸드폰번호",
    "택배수량",
    "택배운임",
    "운임구분",
    "품목명",
    "",
    "배송메세지 (도착일)",
    "보내는 분",
    "연락처",
    "주소",
]


def _parse_iso_to_yyyymmdd(iso_str: str) -> str:
    """ISO-8601 문자열에서 YYYYMMDD 부분만 추출."""
    if not iso_str:
        return ""
    try:
        # '2024-03-01T00:00:00' 또는 '2024-03-01T00:00:00+09:00' 등
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.strftime("%Y%m%d")
    except (ValueError, TypeError):
        return ""


def generate_logen_filename(
    from_iso: Optional[str] = None,
    to_iso: Optional[str] = None,
) -> str:
    """
    Generate filename for Logen shipping Excel file.
    
    기간(from_iso, to_iso)이 둘 다 주어지면 파일명에 반영:
    "로젠발송양식_{시작일}_{종료일}.xlsx"
    기간 미지정 시 오늘 날짜만 사용: "로젠발송양식_{YYYYMMDD}.xlsx"
    
    Args:
        from_iso: 조회 시작 시각 ISO-8601 (선택)
        to_iso: 조회 종료 시각 ISO-8601 (선택)
    
    Returns:
        str: Filename (예: 로젠발송양식_20240301_20240302.xlsx 또는 로젠발송양식_20240302.xlsx)
    """
    if from_iso and to_iso:
        start_str = _parse_iso_to_yyyymmdd(from_iso)
        end_str = _parse_iso_to_yyyymmdd(to_iso)
        if start_str and end_str:
            return f"로젠발송양식_{start_str}_{end_str}.xlsx"
    current_date = datetime.now()
    date_str = current_date.strftime("%Y%m%d")
    return f"로젠발송양식_{date_str}.xlsx"


def read_logen_excel(file_path: str) -> List[Dict[str, str]]:
    """
    Read and verify Logen shipping Excel file.
    
    Reads an Excel file in Logen format (로젠양식: 주소1,2로분리),
    verifies the header row matches the expected column order,
    and parses data rows starting from row 2.
    
    Args:
        file_path: Path to the Excel file to read
        
    Returns:
        List of dictionaries with keys: receiver_name, address1, address2,
        full_address, receiver_tel, product_name, delivery_memo
        
    Raises:
        ValueError: If header row doesn't match expected format
        FileNotFoundError: If file doesn't exist
        
    Example:
        >>> data = read_logen_excel("로젠발송양식_20240115.xlsx")
        >>> data[0]['receiver_name']
        '홍길동'
    """
    # Load the workbook
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    # Parse and verify header row (row 1)
    # openpyxl stores empty cells as None; normalize for comparison
    header_row = sheet[1]
    actual_headers = [(c.value if c.value is not None else "") for c in header_row]
    
    if actual_headers != LOGEN_EXCEL_HEADERS:
        raise ValueError(
            f"Header row mismatch. Expected {LOGEN_EXCEL_HEADERS}, "
            f"but got {actual_headers}"
        )
    
    # Parse data rows starting from row 2
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Skip empty rows (row needs at least 10 columns for valid data)
        row_list = list(row) if row else []
        if len(row_list) < 10 or not any(
            v for v in (row_list[:4] + row_list[8:10]) if v is not None
        ):
            continue
        
        address1 = row_list[2] if len(row_list) > 2 and row_list[2] is not None else ""
        address2 = row_list[3] if len(row_list) > 3 and row_list[3] is not None else ""
        full_address = f"{address1} {address2}".strip() if (address1 or address2) else ""
        
        row_data = {
            "receiver_name": row_list[0] if len(row_list) > 0 and row_list[0] is not None else "",
            "address1": address1,
            "address2": address2,
            "full_address": full_address,
            "receiver_tel": row_list[4] if len(row_list) > 4 and row_list[4] is not None else "",
            "product_name": row_list[9] if len(row_list) > 9 and row_list[9] is not None else "",
            "delivery_memo": row_list[11] if len(row_list) > 11 and row_list[11] is not None else "",
        }
        data.append(row_data)
    
    return data
