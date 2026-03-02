"""Tests for utility functions."""

import pytest
from datetime import datetime
from unittest.mock import patch
import sys
import os

# Import utils directly without triggering package __init__
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))
from utils import generate_logen_filename


def test_generate_logen_filename_format():
    """Test that filename follows the correct format."""
    filename = generate_logen_filename()
    
    # Check that filename starts with Korean prefix
    assert filename.startswith("로젠발송양식_")
    
    # Check that filename ends with .xlsx
    assert filename.endswith(".xlsx")
    
    # Extract date part (between _ and .xlsx)
    date_part = filename[len("로젠발송양식_"):-len(".xlsx")]
    
    # Check that date part is 8 digits (YYYYMMDD)
    assert len(date_part) == 8
    assert date_part.isdigit()


def test_generate_logen_filename_with_known_date():
    """Test filename generation with a known date."""
    # Mock datetime.now() to return a specific date
    mock_date = datetime(2024, 1, 15, 10, 30, 0)
    
    with patch('utils.datetime') as mock_datetime:
        mock_datetime.now.return_value = mock_date
        mock_datetime.strftime = datetime.strftime
        
        filename = generate_logen_filename()
        
        assert filename == "로젠발송양식_20240115.xlsx"


def test_generate_logen_filename_different_dates():
    """Test filename generation with various dates."""
    test_cases = [
        (datetime(2024, 12, 31, 23, 59, 59), "로젠발송양식_20241231.xlsx"),
        (datetime(2024, 1, 1, 0, 0, 0), "로젠발송양식_20240101.xlsx"),
        (datetime(2024, 6, 15, 12, 0, 0), "로젠발송양식_20240615.xlsx"),
    ]
    
    for mock_date, expected_filename in test_cases:
        with patch('utils.datetime') as mock_datetime:
            mock_datetime.now.return_value = mock_date
            mock_datetime.strftime = datetime.strftime
            
            filename = generate_logen_filename()
            
            assert filename == expected_filename, f"Expected {expected_filename}, got {filename}"



def test_read_logen_excel_with_valid_file(tmp_path):
    """Test reading a valid Logen Excel file (로젠양식 A~O 15열)."""
    from openpyxl import Workbook
    from utils import read_logen_excel
    
    # Create a test Excel file (로젠양식 A~O 15열)
    wb = Workbook()
    ws = wb.active
    headers = [
        "수하인명", "운송장번호(로젠택배)", "수하인주소1", "", "",
        "수하인핸드폰번호", "택배수량", "", "", "품목명", "",
        "배송메세지 (도착일)", "보내는 분", "연락처", "주소"
    ]
    ws.append(headers)
    
    # Add data rows (15 columns: A~O)
    ws.append(["홍길동", "", "서울시 강남구 테헤란로 123", "", "", "010-1234-5678", 1, "", "", "테스트 상품", "", "문 앞에 놓아주세요", "", "", ""])
    ws.append(["김철수", "", "부산시 해운대구 해운대로 456", "", "", "010-9876-5432", 1, "", "", "샘플 제품", "", "", "", "", ""])
    
    test_file = tmp_path / "test_logen.xlsx"
    wb.save(test_file)
    
    data = read_logen_excel(str(test_file))
    
    assert len(data) == 2
    assert data[0]["receiver_name"] == "홍길동"
    assert data[0]["address1"] == "서울시 강남구 테헤란로 123"
    assert data[0]["address2"] == ""
    assert data[0]["full_address"] == "서울시 강남구 테헤란로 123"
    assert data[0]["receiver_tel"] == "010-1234-5678"
    assert data[0]["product_name"] == "테스트 상품"
    assert data[0]["delivery_memo"] == "문 앞에 놓아주세요"
    
    assert data[1]["receiver_name"] == "김철수"
    assert data[1]["address1"] == "부산시 해운대구 해운대로 456"
    assert data[1]["address2"] == ""
    assert data[1]["full_address"] == "부산시 해운대구 해운대로 456"
    assert data[1]["receiver_tel"] == "010-9876-5432"
    assert data[1]["product_name"] == "샘플 제품"
    assert data[1]["delivery_memo"] == ""


def test_read_logen_excel_header_only(tmp_path):
    """Test reading Excel file with only header row (로젠양식 A~O 15열)."""
    from openpyxl import Workbook
    from utils import read_logen_excel
    
    wb = Workbook()
    ws = wb.active
    headers = [
        "수하인명", "운송장번호(로젠택배)", "수하인주소1", "", "",
        "수하인핸드폰번호", "택배수량", "", "", "품목명", "",
        "배송메세지 (도착일)", "보내는 분", "연락처", "주소"
    ]
    ws.append(headers)
    
    test_file = tmp_path / "test_header_only.xlsx"
    wb.save(test_file)
    
    data = read_logen_excel(str(test_file))
    assert len(data) == 0


def test_read_logen_excel_invalid_header(tmp_path):
    """Test reading Excel file with invalid header."""
    from openpyxl import Workbook
    from utils import read_logen_excel
    
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Address", "Phone", "Product", "Memo"])
    ws.append(["홍길동", "서울시", "010-1234-5678", "상품", "메모"])
    
    test_file = tmp_path / "test_invalid_header.xlsx"
    wb.save(test_file)
    
    with pytest.raises(ValueError, match="Header row mismatch"):
        read_logen_excel(str(test_file))


def test_read_logen_excel_with_empty_rows(tmp_path):
    """Test reading Excel file with empty rows (로젠양식 A~O 15열)."""
    from openpyxl import Workbook
    from utils import read_logen_excel
    
    wb = Workbook()
    ws = wb.active
    headers = [
        "수하인명", "운송장번호(로젠택배)", "수하인주소1", "", "",
        "수하인핸드폰번호", "택배수량", "", "", "품목명", "",
        "배송메세지 (도착일)", "보내는 분", "연락처", "주소"
    ]
    ws.append(headers)
    ws.append(["홍길동", "", "서울시", "", "", "010-1234-5678", 1, "", "", "상품", "", "메모", "", "", ""])
    ws.append([None] * 15)
    ws.append(["김철수", "", "부산시", "", "", "010-9876-5432", 1, "", "", "제품", "", "", "", "", ""])
    
    test_file = tmp_path / "test_empty_rows.xlsx"
    wb.save(test_file)
    
    data = read_logen_excel(str(test_file))
    assert len(data) == 2
    assert data[0]["receiver_name"] == "홍길동"
    assert data[1]["receiver_name"] == "김철수"


def test_read_logen_excel_with_none_values(tmp_path):
    """Test reading Excel file with None values in cells (로젠양식 A~O 15열)."""
    from openpyxl import Workbook
    from utils import read_logen_excel
    
    wb = Workbook()
    ws = wb.active
    headers = [
        "수하인명", "운송장번호(로젠택배)", "수하인주소1", "", "",
        "수하인핸드폰번호", "택배수량", "", "", "품목명", "",
        "배송메세지 (도착일)", "보내는 분", "연락처", "주소"
    ]
    ws.append(headers)
    row = ["홍길동", "", "서울시", "", "", None, 1, "", "", "상품", "", None, "", "", ""]
    ws.append(row)
    
    test_file = tmp_path / "test_none_values.xlsx"
    wb.save(test_file)
    
    data = read_logen_excel(str(test_file))
    assert len(data) == 1
    assert data[0]["receiver_name"] == "홍길동"
    assert data[0]["address1"] == "서울시"
    assert data[0]["full_address"] == "서울시"
    assert data[0]["receiver_tel"] == ""
    assert data[0]["product_name"] == "상품"
    assert data[0]["delivery_memo"] == ""
