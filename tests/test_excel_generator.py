"""Unit tests for LogenExcelGenerator."""

import pytest
import os
import tempfile
from openpyxl import load_workbook

from src.excel_generator import LogenExcelGenerator
from src.exceptions import ExcelGenerationError


class TestLogenExcelGenerator:
    """Test suite for LogenExcelGenerator class."""
    
    def test_generate_excel_creates_file(self):
        """Test that generate_excel creates a file at the specified path."""
        data = [
            {
                "receiver_name": "홍길동",
                "address1": "서울시 강남구 테헤란로",
                "address2": "123",
                "receiver_tel": "010-1234-5678",
                "product_name": "테스트 상품",
                "delivery_memo": "문 앞에 놓아주세요"
            }
        ]
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            result_path = LogenExcelGenerator.generate_excel(data, output_path)
            
            assert result_path == output_path
            assert os.path.exists(output_path)
        finally:
            if os.path.exists(output_path):
                os.remove(output_path)
    
    def test_generate_excel_header_row(self):
        """Test that the Excel file has the correct header row (로젠양식 15열)."""
        data = []
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            LogenExcelGenerator.generate_excel(data, output_path)
            
            wb = load_workbook(output_path)
            ws = wb.active
            
            expected_headers = [
                "수하인명", "운송장번호(로젠택배)", "수하인주소1", "수하인주소2",
                "수하인전화번호", "수하인핸드폰번호", "택배수량", "택배운임", "운임구분",
                "품목명", "", "배송메세지 (도착일)", "보내는 분", "연락처", "주소"
            ]
            actual_headers = [(c.value if c.value is not None else "") for c in ws[1]]
            
            assert actual_headers == expected_headers
        finally:
            if os.path.exists(output_path):
                os.remove(output_path)
    
    def test_generate_excel_single_order(self):
        """Test that a single order is correctly written to Excel (로젠양식)."""
        data = [
            {
                "receiver_name": "김철수",
                "address1": "부산시 해운대구 센텀로",
                "address2": "456",
                "receiver_tel": "010-9876-5432",
                "product_name": "노트북",
                "delivery_memo": "배송 전 연락주세요"
            }
        ]
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            LogenExcelGenerator.generate_excel(data, output_path)
            
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Check row count (1 header + 1 data row)
            assert ws.max_row == 2
            
            # Check data in row 2 (로젠양식 열 순서)
            assert ws.cell(2, 1).value == "김철수"  # 수하인명
            assert ws.cell(2, 2).value in (None, "")  # 운송장번호 (빈칸)
            assert ws.cell(2, 3).value == "부산시 해운대구 센텀로"  # 수하인주소1
            assert ws.cell(2, 4).value == "456"  # 수하인주소2
            assert ws.cell(2, 5).value == "010-9876-5432"  # 수하인전화번호
            assert ws.cell(2, 10).value == "노트북"  # 품목명
            assert ws.cell(2, 12).value == "배송 전 연락주세요"  # 배송메세지
        finally:
            if os.path.exists(output_path):
                os.remove(output_path)
    
    def test_generate_excel_multiple_orders(self):
        """Test that multiple orders are correctly written to Excel (로젠양식)."""
        data = [
            {
                "receiver_name": "이영희",
                "address1": "대구시 중구 동성로",
                "address2": "111",
                "receiver_tel": "010-1111-2222",
                "product_name": "키보드",
                "delivery_memo": ""
            },
            {
                "receiver_name": "박민수",
                "address1": "인천시 남동구 구월로",
                "address2": "222",
                "receiver_tel": "010-3333-4444",
                "product_name": "마우스",
                "delivery_memo": "부재시 경비실"
            },
            {
                "receiver_name": "최지훈",
                "address1": "광주시 서구 상무대로",
                "address2": "333",
                "receiver_tel": "010-5555-6666",
                "product_name": "모니터",
                "delivery_memo": "조심히 다뤄주세요"
            }
        ]
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            LogenExcelGenerator.generate_excel(data, output_path)
            
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Check row count (1 header + 3 data rows)
            assert ws.max_row == 4
            
            # Check first data row (품목명=열10)
            assert ws.cell(2, 1).value == "이영희"
            assert ws.cell(2, 10).value == "키보드"
            
            # Check second data row
            assert ws.cell(3, 1).value == "박민수"
            assert ws.cell(3, 10).value == "마우스"
            
            # Check third data row
            assert ws.cell(4, 1).value == "최지훈"
            assert ws.cell(4, 10).value == "모니터"
        finally:
            if os.path.exists(output_path):
                os.remove(output_path)
    
    def test_generate_excel_empty_data(self):
        """Test that empty data list creates only header row."""
        data = []
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            LogenExcelGenerator.generate_excel(data, output_path)
            
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Should have only header row
            assert ws.max_row == 1
        finally:
            if os.path.exists(output_path):
                os.remove(output_path)
    
    def test_generate_excel_empty_delivery_memo(self):
        """Test that empty delivery_memo is handled correctly."""
        data = [
            {
                "receiver_name": "정수진",
                "address1": "대전시 유성구 대학로",
                "address2": "444",
                "receiver_tel": "010-7777-8888",
                "product_name": "헤드폰",
                "delivery_memo": ""
            }
        ]
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            LogenExcelGenerator.generate_excel(data, output_path)
            
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Check that delivery_memo (열12) is None or empty
            assert ws.cell(2, 12).value is None or ws.cell(2, 12).value == ""
        finally:
            if os.path.exists(output_path):
                os.remove(output_path)
    
    def test_generate_excel_missing_fields(self):
        """Test that missing fields are handled with empty strings (full_address fallback)."""
        data = [
            {
                "receiver_name": "강민지",
                "full_address": "울산시 남구 삼산로 555"
                # Missing address1, address2, receiver_tel, product_name, delivery_memo
            }
        ]
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            LogenExcelGenerator.generate_excel(data, output_path)
            
            wb = load_workbook(output_path)
            ws = wb.active
            
            # full_address fallback: address1 gets full, address2 empty
            assert ws.cell(2, 1).value == "강민지"
            assert ws.cell(2, 3).value == "울산시 남구 삼산로 555"  # address1
            assert ws.cell(2, 4).value is None or ws.cell(2, 4).value == ""  # address2
            assert ws.cell(2, 5).value is None or ws.cell(2, 5).value == ""
            assert ws.cell(2, 10).value is None or ws.cell(2, 10).value == ""
        finally:
            if os.path.exists(output_path):
                os.remove(output_path)
    
    def test_generate_excel_invalid_path_raises_error(self):
        """Test that invalid file path raises ExcelGenerationError."""
        data = [
            {
                "receiver_name": "테스트",
                "address1": "주소",
                "address2": "",
                "receiver_tel": "010-0000-0000",
                "product_name": "상품",
                "delivery_memo": ""
            }
        ]
        
        # Use an invalid path (directory that doesn't exist)
        invalid_path = "/nonexistent/directory/file.xlsx"
        
        with pytest.raises(ExcelGenerationError) as exc_info:
            LogenExcelGenerator.generate_excel(data, invalid_path)
        
        assert exc_info.value.file_path == invalid_path
        assert exc_info.value.underlying_error is not None
    
    def test_generate_excel_column_order(self):
        """Test that columns are in the correct order (로젠양식)."""
        data = [
            {
                "receiver_name": "A",
                "address1": "B1",
                "address2": "B2",
                "receiver_tel": "C",
                "product_name": "D",
                "delivery_memo": "E"
            }
        ]
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            LogenExcelGenerator.generate_excel(data, output_path)
            
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Check column order (로젠양식 15열)
            assert ws.cell(2, 1).value == "A"   # 수하인명
            assert ws.cell(2, 3).value == "B1"  # 수하인주소1
            assert ws.cell(2, 4).value == "B2"  # 수하인주소2
            assert ws.cell(2, 5).value == "C"   # 수하인전화번호
            assert ws.cell(2, 10).value == "D"  # 품목명
            assert ws.cell(2, 12).value == "E"  # 배송메세지
        finally:
            if os.path.exists(output_path):
                os.remove(output_path)
